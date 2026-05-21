#!/usr/bin/env python3
"""
OpenFOAM Post-Processing Plotter
=================================
Recursively finds every 'postProcessing' folder under a root directory and:

  Per postProcessing folder:
    - forceCoeffs.png          Cd, Cl, Cm (raw)
    - yPlus.png                all patches, all stats combined
    - yPlus_min.png            min y+ for each patch
    - yPlus_max.png            max y+ for each patch
    - yPlus_mean.png           mean y+ for each patch

  Per mesh folder (e.g. coarse_mesh/):
    - forceCoeffs_combined.png   all timesteps overlaid
    - yPlus_combined.png         all patches + stats, all timesteps overlaid
    - yPlus_min_combined.png     min only, all timesteps overlaid
    - yPlus_max_combined.png     max only, all timesteps overlaid
    - yPlus_mean_combined.png    mean only, all timesteps overlaid
    - stats.xlsx                 min/max/mean summary for every timestep

Timestep naming: ts001 → dt=0.001, ts0001 → dt=0.0001, etc.

Usage:
    python plot_openfoam.py [root_dir]
Defaults to current working directory.
"""

import os
import sys
import re
import glob
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────

PATCHES = {
    "cube": {"label": "Cube", "color": "#ffa657", "fill": "#3d1f05"},
}
PATCH_ORDER  = ["cube"]
EXTRA_COLORS = ["#d2a8ff", "#f778ba", "#79c0ff", "#56d364"]
FC_COLORS    = {"Cd": "#58a6ff", "Cl": "#3fb950", "Cm": "#f78166"}
TS_PALETTE   = [
    "#58a6ff", "#3fb950", "#ffa657", "#f78166",
    "#d2a8ff", "#f778ba", "#79c0ff", "#56d364",
]

STAT_TITLES = {
    "min":  "Minimum y⁺",
    "max":  "Maximum y⁺",
    "mean": "Mean y⁺",
}


# ──────────────────────────────────────────────
# TIMESTEP NAME → FLOAT
# ──────────────────────────────────────────────

def ts_to_float(ts_name):
    stem = re.sub(r"^ts", "", ts_name.lower())
    if not stem:
        return None
    try:
        return float("0." + stem)
    except ValueError:
        return None


def ts_label(ts_name):
    val = ts_to_float(ts_name)
    return f"dt = {val:g}" if val is not None else ts_name


# ──────────────────────────────────────────────
# FOLDER DISCOVERY
# ──────────────────────────────────────────────

def find_postprocessing_dirs(root):
    matches = []
    for dirpath, dirnames, _ in os.walk(root):
        for d in dirnames:
            if d.lower() == "postprocessing":
                matches.append(os.path.join(dirpath, d))
    return sorted(matches)


def find_files(base_dir, pattern):
    return sorted(glob.glob(os.path.join(base_dir, "**", pattern), recursive=True))


def relative_label(path, root):
    try:
        return " / ".join(Path(path).relative_to(root).parts)
    except ValueError:
        return str(path)


def infer_mesh_dir(pp_dir):
    return str(Path(pp_dir).parent.parent)


def infer_ts_name(pp_dir):
    return Path(pp_dir).parent.name


def group_by_mesh(pp_dirs):
    groups = defaultdict(list)
    for pp_dir in pp_dirs:
        groups[infer_mesh_dir(pp_dir)].append((infer_ts_name(pp_dir), pp_dir))
    for mesh_dir in groups:
        groups[mesh_dir].sort(key=lambda x: ts_to_float(x[0]) or 0)
    return dict(groups)


# ──────────────────────────────────────────────
# PARSERS
# ──────────────────────────────────────────────

def parse_force_coeffs(filepath):
    data = {"time": [], "Cd": [], "Cl": [], "Cm": []}
    col_map = None
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("#"):
                low = line.lower()
                if "cd" in low or "cl" in low:
                    parts = line.lstrip("#").split()
                    col_map = {n.lower(): i for i, n in enumerate(parts)}
                continue
            parts = line.split()
            try:
                row = [float(x) for x in parts]
            except ValueError:
                continue
            if col_map:
                ti = col_map.get("time", 0)
                di = col_map.get("cd",   1)
                li = col_map.get("cl",   3)
                mi = col_map.get("cm",   None)
            else:
                ti, mi, di, li = 0, 1, 2, 3
            needed = [x for x in [ti, di, li] if x is not None]
            if len(row) <= max(needed):
                continue
            data["time"].append(row[ti])
            data["Cd"].append(row[di])
            data["Cl"].append(row[li])
            if mi is not None and len(row) > mi:
                data["Cm"].append(row[mi])
    return {k: np.array(v) for k, v in data.items()}


def parse_yplus(filepath):
    patch_data = {}
    with open(filepath) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) >= 5:
                try:
                    t = float(parts[0])
                    float(parts[1])
                except ValueError:
                    try:
                        pk = parts[1].lower()
                        mn = float(parts[2])
                        mx = float(parts[3])
                        av = float(parts[4])
                    except (ValueError, IndexError):
                        continue
                    if pk not in patch_data:
                        patch_data[pk] = {"time": [], "min": [], "max": [], "mean": []}
                    patch_data[pk]["time"].append(t)
                    patch_data[pk]["min"].append(mn)
                    patch_data[pk]["max"].append(mx)
                    patch_data[pk]["mean"].append(av)
                    continue
            if len(parts) >= 4:
                try:
                    row = [float(x) for x in parts[:4]]
                except ValueError:
                    continue
                if None not in patch_data:
                    patch_data[None] = {"time": [], "min": [], "max": [], "mean": []}
                patch_data[None]["time"].append(row[0])
                patch_data[None]["min"].append(row[1])
                patch_data[None]["max"].append(row[2])
                patch_data[None]["mean"].append(row[3])
    return {k: {s: np.array(v) for s, v in d.items()} for k, d in patch_data.items()}


def load_pp_data(pp_dir):
    fc_files = (find_files(pp_dir, "forceCoeffs.dat")
                or find_files(pp_dir, "coefficient.dat")
                or find_files(pp_dir, "forceCoeffs_*.dat"))
    all_yp = sorted(set(
        find_files(pp_dir, "yPlus.dat") + find_files(pp_dir, "yPlus_*.dat")
    ))
    fc_list = []
    for fp in fc_files:
        try:
            fc = parse_force_coeffs(fp)
            if len(fc["time"]):
                fc_list.append(fc)
        except Exception:
            pass
    patch_data = {}
    for fp in all_yp:
        try:
            result = parse_yplus(fp)
        except Exception:
            continue
        for pk, arrays in result.items():
            if len(arrays["time"]) == 0:
                continue
            if pk is None:
                pk = next(
                    (key for part in Path(fp).parts for key in PATCHES
                     if key in part.lower()),
                    Path(fp).parent.parent.name.lower()
                )
            patch_data[pk] = arrays
    return fc_list, patch_data


def ordered_patches(patch_data):
    # Only include patches listed in PATCH_ORDER — ignores all others
    return [k for k in PATCH_ORDER if k in patch_data]



def patch_cfg(pk, extra_idx=0):
    if pk in PATCHES:
        return PATCHES[pk], extra_idx
    c = EXTRA_COLORS[extra_idx % len(EXTRA_COLORS)]
    return {"label": pk, "color": c, "fill": c + "22"}, extra_idx + 1


# ──────────────────────────────────────────────
# STYLE
# ──────────────────────────────────────────────

def setup_style():
    plt.rcParams.update({
        "figure.facecolor": "#0d1117",
        "axes.facecolor":   "#161b22",
        "axes.edgecolor":   "#30363d",
        "axes.labelcolor":  "#e6edf3",
        "axes.titlecolor":  "#e6edf3",
        "xtick.color":      "#8b949e",
        "ytick.color":      "#8b949e",
        "grid.color":       "#21262d",
        "grid.linewidth":   0.8,
        "text.color":       "#e6edf3",
        "legend.facecolor": "#161b22",
        "legend.edgecolor": "#30363d",
        "font.family":      "monospace",
        "font.size":        10,
    })


def save_fig(fig, path):
    fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


# ──────────────────────────────────────────────
# PLOT PRIMITIVES
# ──────────────────────────────────────────────

def _ref_lines(ax, ymax, t):
    """Draw y⁺ reference lines."""
    for ref, a, lbl in [(1, 0.85, "y⁺=1"), (5, 0.6, "y⁺=5"),
                        (30, 0.45, "y⁺=30"), (300, 0.35, "y⁺=300")]:
        if ref < ymax * 1.5:
            ax.axhline(ref, color="#8b949e", lw=0.9, ls=":", alpha=a)
            if len(t):
                ax.text(t[-1], ref * 1.06, lbl, color="#8b949e",
                        fontsize=7, va="bottom", ha="right")


def plot_fc_on_axes(fc, ax_cd, ax_cl, ax_cm, color=None, label=""):
    t, lw = fc["time"], 1.4
    for ax, arr, key, ylabel, col in [
        (ax_cd, fc["Cd"], "Cd", "C$_d$", color or FC_COLORS["Cd"]),
        (ax_cl, fc["Cl"], "Cl", "C$_l$", color or FC_COLORS["Cl"]),
    ]:
        ax.plot(t, arr, color=col, linewidth=lw, label=label or key)
        ax.set_ylabel(ylabel)
        ax.grid(True, linestyle="--", alpha=0.4)
    cm = fc["Cm"]
    if len(cm):
        ax_cm.plot(t, cm, color=color or FC_COLORS["Cm"], linewidth=lw,
                   label=label or "Cm")
    ax_cm.set_ylabel("C$_m$")
    ax_cm.grid(True, linestyle="--", alpha=0.4)


def slice_last_second(fc):
    """Return a copy of fc dict sliced to the final 1 second of data."""
    t = fc.get("time", np.array([]))
    if len(t) == 0:
        return fc
    mask = t >= (t[-1] - 1.0)
    return {k: v[mask] if isinstance(v, np.ndarray) and len(v) == len(t) else v
            for k, v in fc.items()}


def slice_last_second_yp(yp):
    """Return a copy of yp dict sliced to the final 1 second of data."""
    t = yp.get("time", np.array([]))
    if len(t) == 0:
        return yp
    mask = t >= (t[-1] - 1.0)
    return {k: v[mask] if isinstance(v, np.ndarray) and len(v) == len(t) else v
            for k, v in yp.items()}


def plot_yplus_combined_ax(yp, ax, cfg, label_suffix="", alpha_fill=1.0, lw=2.0):
    """All three stats (min/max/mean bands) on one axis."""
    t, mn, mx, av = yp["time"], yp["min"], yp["max"], yp["mean"]
    color = cfg["color"]
    ax.fill_between(t, mn, mx, color=cfg["fill"], alpha=alpha_fill, zorder=1)
    ax.plot(t, mn, color=color, lw=1.0, ls="--", alpha=0.65, zorder=2,
            label=f"min {label_suffix}".strip())
    ax.plot(t, mx, color=color, lw=1.0, ls=":",  alpha=0.65, zorder=2,
            label=f"max {label_suffix}".strip())
    ax.plot(t, av, color=color, lw=lw, zorder=3,
            label=f"mean {label_suffix}".strip())
    _ref_lines(ax, mx.max(), t)
    ax.set_ylabel("y$^+$")
    ax.set_yscale("linear")
    ax.grid(True, which="both", linestyle="--", alpha=0.35)
    ax.legend(fontsize=7, loc="upper right")


def plot_yplus_stat_ax(yp, ax, cfg, stat, label_suffix="", lw=2.0):
    """Single stat ('min', 'max', or 'mean') on one axis."""
    t   = yp["time"]
    arr = yp[stat]
    ax.plot(t, arr, color=cfg["color"], lw=lw, zorder=2,
            label=label_suffix.strip() or cfg["label"])
    _ref_lines(ax, arr.max(), t)
    ax.set_ylabel("y$^+$")
    ax.set_yscale("linear")
    ax.grid(True, which="both", linestyle="--", alpha=0.35)
    ax.legend(fontsize=8, loc="upper right")


# ──────────────────────────────────────────────
# INDIVIDUAL PP FOLDER PLOTS
# ──────────────────────────────────────────────

def _make_yplus_fig(patch_data, label, title_prefix, plot_fn, filename, out_dir):
    """Generic helper: one subplot per patch, using plot_fn(yp, ax, cfg)."""
    patches = ordered_patches(patch_data)
    n = len(patches)
    fig = plt.figure(figsize=(13, 4.0 * n), constrained_layout=True)
    fig.suptitle(f"{title_prefix}\n{label}", fontsize=12,
                 fontweight="bold", color="#e6edf3")
    gs = gridspec.GridSpec(n, 1, figure=fig)
    extra_idx = 0
    for i, pk in enumerate(patches):
        ax = fig.add_subplot(gs[i])
        cfg, extra_idx = patch_cfg(pk, extra_idx)
        plot_fn(patch_data[pk], ax, cfg)
        ax.set_title(f"y$^+$  —  {cfg['label']}", fontsize=11, loc="left", pad=5)
        ax.set_xlabel("Time (s)")
    save_fig(fig, os.path.join(out_dir, filename))


def save_individual_plots(pp_dir, root):
    label = relative_label(pp_dir, root)
    fc_list, patch_data = load_pp_data(pp_dir)
    fc_saved = yp_saved = False

    # Force coefficients
    if fc_list:
        fig = plt.figure(figsize=(13, 9), constrained_layout=True)
        fig.suptitle(f"Force Coefficients\n{label}", fontsize=12,
                     fontweight="bold", color="#e6edf3")
        gs = gridspec.GridSpec(3, 1, figure=fig)
        ax_cd, ax_cl, ax_cm = [fig.add_subplot(gs[i]) for i in range(3)]
        for fc in fc_list:
            plot_fc_on_axes(fc, ax_cd, ax_cl, ax_cm)
        for ax in [ax_cd, ax_cl, ax_cm]:
            ax.legend(fontsize=9)
        ax_cm.set_xlabel("Time (s)")
        save_fig(fig, os.path.join(pp_dir, "forceCoeffs.png"))

        # Last-second forceCoeffs plot
        fig2 = plt.figure(figsize=(13, 9), constrained_layout=True)
        fig2.suptitle(f"Force Coefficients — Last 1 s\n{label}", fontsize=12,
                      fontweight="bold", color="#e6edf3")
        gs2 = gridspec.GridSpec(3, 1, figure=fig2)
        ax2_cd, ax2_cl, ax2_cm = [fig2.add_subplot(gs2[i]) for i in range(3)]
        for fc in fc_list:
            plot_fc_on_axes(slice_last_second(fc), ax2_cd, ax2_cl, ax2_cm)
        for ax in [ax2_cd, ax2_cl, ax2_cm]:
            ax.legend(fontsize=9)
        ax2_cm.set_xlabel("Time (s)")
        save_fig(fig2, os.path.join(pp_dir, "forceCoeffs_last1s.png"))
        fc_saved = True

    # yPlus figures
    if patch_data:
        # Combined (all stats together)
        _make_yplus_fig(
            patch_data, label, "Wall y⁺",
            lambda yp, ax, cfg: plot_yplus_combined_ax(yp, ax, cfg),
            "yPlus.png", pp_dir
        )
        # Separate stat images
        for stat in ("min", "max", "mean"):
            _make_yplus_fig(
                patch_data, label, STAT_TITLES[stat],
                lambda yp, ax, cfg, s=stat: plot_yplus_stat_ax(yp, ax, cfg, s),
                f"yPlus_{stat}.png", pp_dir
            )
        # Last-second yPlus plots
        def _make_yplus_last1s(pd_full, label, title_prefix, plot_fn, filename, out_dir):
            pd_sliced = {pk: slice_last_second_yp(yp) for pk, yp in pd_full.items()}
            _make_yplus_fig(pd_sliced, label, title_prefix, plot_fn, filename, out_dir)

        _make_yplus_last1s(
            patch_data, label, "Wall y⁺ — Last 1 s",
            lambda yp, ax, cfg: plot_yplus_combined_ax(yp, ax, cfg),
            "yPlus_last1s.png", pp_dir
        )
        for stat in ("min", "max", "mean"):
            _make_yplus_last1s(
                patch_data, label, f"{STAT_TITLES[stat]} — Last 1 s",
                lambda yp, ax, cfg, s=stat: plot_yplus_stat_ax(yp, ax, cfg, s),
                f"yPlus_{stat}_last1s.png", pp_dir
            )
        yp_saved = True

    return fc_saved, yp_saved


# ──────────────────────────────────────────────
# COMBINED MESH-LEVEL PLOTS
# ──────────────────────────────────────────────

def _collect_ts_data(ts_entries):
    ts_data = []
    for ts_name, pp_dir in ts_entries:
        fc_list, patch_data = load_pp_data(pp_dir)
        ts_data.append({
            "ts":         ts_name,
            "label":      ts_label(ts_name),
            "fc_list":    fc_list,
            "patch_data": patch_data,
        })
    return ts_data


def _all_patches(ts_data):
    seen = []
    for d in ts_data:
        for pk in d["patch_data"]:
            if pk not in seen:
                seen.append(pk)
    # Only include patches listed in PATCH_ORDER — ignores all others
    return [k for k in PATCH_ORDER if k in seen]



def save_combined_plots(mesh_dir, ts_entries, root):
    mesh_label = relative_label(mesh_dir, root)
    ts_data    = _collect_ts_data(ts_entries)
    patches    = _all_patches(ts_data)

    # ── combined forceCoeffs ──────────────────
    if any(d["fc_list"] for d in ts_data):
        fig = plt.figure(figsize=(13, 9), constrained_layout=True)
        fig.suptitle(f"Force Coefficients — Combined Timesteps\n{mesh_label}",
                     fontsize=12, fontweight="bold", color="#e6edf3")
        gs = gridspec.GridSpec(3, 1, figure=fig)
        ax_cd, ax_cl, ax_cm = [fig.add_subplot(gs[i]) for i in range(3)]
        for i, d in enumerate(ts_data):
            color = TS_PALETTE[i % len(TS_PALETTE)]
            for fc in d["fc_list"]:
                plot_fc_on_axes(fc, ax_cd, ax_cl, ax_cm,
                                color=color, label=d["label"])
        for ax in [ax_cd, ax_cl, ax_cm]:
            ax.legend(fontsize=9)
        ax_cm.set_xlabel("Time (s)")
        save_fig(fig, os.path.join(mesh_dir, "forceCoeffs_combined.png"))

        # Last-second combined forceCoeffs
        fig = plt.figure(figsize=(13, 9), constrained_layout=True)
        fig.suptitle(f"Force Coefficients — Last 1 s — Combined Timesteps\n{mesh_label}",
                     fontsize=12, fontweight="bold", color="#e6edf3")
        gs = gridspec.GridSpec(3, 1, figure=fig)
        ax_cd, ax_cl, ax_cm = [fig.add_subplot(gs[i]) for i in range(3)]
        for i, d in enumerate(ts_data):
            color = TS_PALETTE[i % len(TS_PALETTE)]
            for fc in d["fc_list"]:
                plot_fc_on_axes(slice_last_second(fc), ax_cd, ax_cl, ax_cm,
                                color=color, label=d["label"])
        for ax in [ax_cd, ax_cl, ax_cm]:
            ax.legend(fontsize=9)
        ax_cm.set_xlabel("Time (s)")
        save_fig(fig, os.path.join(mesh_dir, "forceCoeffs_last1s_combined.png"))

    if not patches:
        return

    # ── combined yPlus (all stats) ────────────
    n = len(patches)
    fig = plt.figure(figsize=(13, 4.5 * n), constrained_layout=True)
    fig.suptitle(f"Wall y⁺ — Combined Timesteps\n{mesh_label}",
                 fontsize=12, fontweight="bold", color="#e6edf3")
    gs = gridspec.GridSpec(n, 1, figure=fig)
    for i, pk in enumerate(patches):
        ax = fig.add_subplot(gs[i])
        cfg_base = PATCHES.get(pk, {"label": pk,
                                    "color": EXTRA_COLORS[0],
                                    "fill":  EXTRA_COLORS[0] + "22"})
        ax.set_title(f"y$^+$  —  {cfg_base['label']}", fontsize=11, loc="left", pad=5)
        for j, d in enumerate(ts_data):
            if pk not in d["patch_data"]:
                continue
            color = TS_PALETTE[j % len(TS_PALETTE)]
            cfg   = {"label": cfg_base["label"], "color": color, "fill": color + "18"}
            plot_yplus_combined_ax(d["patch_data"][pk], ax, cfg,
                                   label_suffix=d["label"],
                                   alpha_fill=0.25 if len(ts_data) > 1 else 1.0,
                                   lw=1.8)
        ax.set_xlabel("Time (s)")
    save_fig(fig, os.path.join(mesh_dir, "yPlus_combined.png"))

    # ── separate stat combined plots ──────────
    for stat in ("min", "max", "mean"):
        fig = plt.figure(figsize=(13, 4.0 * n), constrained_layout=True)
        fig.suptitle(f"{STAT_TITLES[stat]} — Combined Timesteps\n{mesh_label}",
                     fontsize=12, fontweight="bold", color="#e6edf3")
        gs = gridspec.GridSpec(n, 1, figure=fig)
        for i, pk in enumerate(patches):
            ax = fig.add_subplot(gs[i])
            cfg_base = PATCHES.get(pk, {"label": pk,
                                        "color": EXTRA_COLORS[0],
                                        "fill":  EXTRA_COLORS[0] + "22"})
            ax.set_title(f"{STAT_TITLES[stat]}  —  {cfg_base['label']}",
                         fontsize=11, loc="left", pad=5)
            for j, d in enumerate(ts_data):
                if pk not in d["patch_data"]:
                    continue
                color = TS_PALETTE[j % len(TS_PALETTE)]
                cfg   = {"label": cfg_base["label"], "color": color,
                         "fill": color + "18"}
                plot_yplus_stat_ax(d["patch_data"][pk], ax, cfg,
                                   stat, label_suffix=d["label"], lw=1.8)
            ax.set_xlabel("Time (s)")
        save_fig(fig, os.path.join(mesh_dir, f"yPlus_{stat}_combined.png"))

    # ── last-second combined yPlus ────────────
    fig = plt.figure(figsize=(13, 4.5 * n), constrained_layout=True)
    fig.suptitle(f"Wall y⁺ — Last 1 s — Combined Timesteps\n{mesh_label}",
                 fontsize=12, fontweight="bold", color="#e6edf3")
    gs = gridspec.GridSpec(n, 1, figure=fig)
    for i, pk in enumerate(patches):
        ax = fig.add_subplot(gs[i])
        cfg_base = PATCHES.get(pk, {"label": pk,
                                    "color": EXTRA_COLORS[0],
                                    "fill":  EXTRA_COLORS[0] + "22"})
        ax.set_title(f"y$^+$ — Last 1 s  —  {cfg_base['label']}", fontsize=11,
                     loc="left", pad=5)
        for j, d in enumerate(ts_data):
            if pk not in d["patch_data"]:
                continue
            color = TS_PALETTE[j % len(TS_PALETTE)]
            cfg   = {"label": cfg_base["label"], "color": color, "fill": color + "18"}
            plot_yplus_combined_ax(slice_last_second_yp(d["patch_data"][pk]),
                                   ax, cfg, label_suffix=d["label"],
                                   alpha_fill=0.25 if len(ts_data) > 1 else 1.0,
                                   lw=1.8)
        ax.set_xlabel("Time (s)")
    save_fig(fig, os.path.join(mesh_dir, "yPlus_last1s_combined.png"))

    # ── last-second separate stat combined plots ──
    for stat in ("min", "max", "mean"):
        fig = plt.figure(figsize=(13, 4.0 * n), constrained_layout=True)
        fig.suptitle(f"{STAT_TITLES[stat]} — Last 1 s — Combined Timesteps\n{mesh_label}",
                     fontsize=12, fontweight="bold", color="#e6edf3")
        gs = gridspec.GridSpec(n, 1, figure=fig)
        for i, pk in enumerate(patches):
            ax = fig.add_subplot(gs[i])
            cfg_base = PATCHES.get(pk, {"label": pk,
                                        "color": EXTRA_COLORS[0],
                                        "fill":  EXTRA_COLORS[0] + "22"})
            ax.set_title(f"{STAT_TITLES[stat]} — Last 1 s  —  {cfg_base['label']}",
                         fontsize=11, loc="left", pad=5)
            for j, d in enumerate(ts_data):
                if pk not in d["patch_data"]:
                    continue
                color = TS_PALETTE[j % len(TS_PALETTE)]
                cfg   = {"label": cfg_base["label"], "color": color, "fill": color + "18"}
                plot_yplus_stat_ax(slice_last_second_yp(d["patch_data"][pk]),
                                   ax, cfg, stat, label_suffix=d["label"], lw=1.8)
            ax.set_xlabel("Time (s)")
        save_fig(fig, os.path.join(mesh_dir, f"yPlus_{stat}_last1s_combined.png"))


# ──────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────

HDR_FILL  = PatternFill("solid", start_color="1F3864")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUB_FILL  = PatternFill("solid", start_color="2E4057")
SUB_FONT  = Font(bold=True, color="D0E8FF", name="Arial", size=10)
CELL_FONT = Font(name="Arial", size=10)
ALT_FILL  = PatternFill("solid", start_color="1A2332")
_thin     = Side(style="thin", color="30363D")
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER    = Alignment(horizontal="center", vertical="center")
RIGHT     = Alignment(horizontal="right")
NUM_FMT   = "0.0000"


def _hdr(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = CENTER; c.border = BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _sub(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = SUB_FONT; c.fill = SUB_FILL
    c.alignment = CENTER; c.border = BORDER
    return c


def _val(ws, row, col, value, alt=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = CELL_FONT; c.border = BORDER
    if isinstance(value, float):
        c.number_format = NUM_FMT; c.alignment = RIGHT
    if alt:
        c.fill = ALT_FILL
    return c


def write_excel(mesh_dir, ts_entries, root):
    wb = Workbook()
    wb.remove(wb.active)
    mesh_label = relative_label(mesh_dir, root)
    ts_data    = _collect_ts_data(ts_entries)

    # ── yPlus sheets ──────────────────────────
    all_pk = _all_patches(ts_data)
    for pk in all_pk:
        patch_label = PATCHES.get(pk, {}).get("label", pk)
        ws = wb.create_sheet(title=f"yPlus_{patch_label[:20]}")

        # Title
        ncols = 1 + len(ts_data) * 4
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        tc = ws.cell(row=1, column=1,
                     value=f"y⁺ Statistics — {patch_label}  |  {mesh_label}")
        tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        tc.fill = PatternFill("solid", start_color="0D1117")
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 22

        _hdr(ws, 2, 1, "Time (s)", width=12)
        _hdr(ws, 3, 1, "")

        col = 2
        for d in ts_data:
            _hdr(ws, 2, col, d["label"], width=14)
            ws.merge_cells(start_row=2, start_column=col,
                           end_row=2, end_column=col + 2)
            for j, stat in enumerate(["min", "max", "mean"]):
                _sub(ws, 3, col + j, stat)
                ws.column_dimensions[get_column_letter(col + j)].width = 13
            col += 3

        # Time axis from the timestep with the most steps
        ref_d = max(
            (d for d in ts_data if pk in d["patch_data"]),
            key=lambda d: len(d["patch_data"][pk]["time"]),
            default=None
        )
        if ref_d is None:
            continue
        times = ref_d["patch_data"][pk]["time"]

        for i, t in enumerate(times):
            r, alt = 4 + i, i % 2 == 1
            _val(ws, r, 1, float(t), alt)
            col = 2
            for d in ts_data:
                yp = d["patch_data"].get(pk)
                if yp is not None and i < len(yp["time"]):
                    _val(ws, r, col,     float(yp["min"][i]),  alt)
                    _val(ws, r, col + 1, float(yp["max"][i]),  alt)
                    _val(ws, r, col + 2, float(yp["mean"][i]), alt)
                else:
                    for j in range(3):
                        _val(ws, r, col + j, "N/A", alt)
                col += 3

        # Summary rows
        n_data = len(times)
        for stat_lbl, fn_name in [("Overall Min",  "MIN"),
                                   ("Overall Max",  "MAX"),
                                   ("Overall Mean", "AVERAGE")]:
            r = 4 + n_data + 1
            c = ws.cell(row=r, column=1, value=stat_lbl)
            c.font = Font(bold=True, name="Arial", size=10, color="FFA657")
            c.border = BORDER
            col = 2
            for d in ts_data:
                for j in range(3):
                    dc   = get_column_letter(col + j)
                    rng  = f"{dc}4:{dc}{3 + n_data}"
                    cell = ws.cell(row=r, column=col + j,
                                   value=f"={fn_name}({rng})")
                    cell.font          = Font(bold=True, name="Arial",
                                              size=10, color="FFA657")
                    cell.number_format = NUM_FMT
                    cell.border        = BORDER
                    cell.alignment     = RIGHT
                col += 3

        ws.freeze_panes = "B4"

    # ── forceCoeffs sheet ─────────────────────
    if any(d["fc_list"] for d in ts_data):
        ws = wb.create_sheet(title="forceCoeffs")
        ncols = 1 + len(ts_data) * 4
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        tc = ws.cell(row=1, column=1,
                     value=f"Force Coefficient Statistics  |  {mesh_label}")
        tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
        tc.fill = PatternFill("solid", start_color="0D1117")
        tc.alignment = CENTER
        ws.row_dimensions[1].height = 22

        _hdr(ws, 2, 1, "Coefficient", width=14)
        _hdr(ws, 3, 1, "Statistic",   width=22)
        col = 2
        for d in ts_data:
            _hdr(ws, 2, col, d["label"], width=14)
            ws.merge_cells(start_row=2, start_column=col,
                           end_row=2, end_column=col + 3)
            for j, stat in enumerate(["min", "max", "mean", "final"]):
                _sub(ws, 3, col + j, stat)
                ws.column_dimensions[get_column_letter(col + j)].width = 13
            col += 4

        HIGHLIGHT_GREEN  = "56D364"
        HIGHLIGHT_YELLOW = "E3B341"

        def _last_second_mean(fc, key):
            """Mean of all samples in the final 1 second of simulation time."""
            t   = fc.get("time", np.array([]))
            arr = fc.get(key,    np.array([]))
            if len(t) == 0:
                return None
            t_end   = t[-1]
            t_start = t_end - 1.0
            mask    = t >= t_start
            subset  = arr[mask]
            return float(subset.mean()) if len(subset) else float(arr[-1])

        r = 4
        for coeff_name, key in [("Cd", "Cd"), ("Cl", "Cl"), ("Cm", "Cm")]:
            rows = [
                ("min",                   np.min,  None),
                ("max",                   np.max,  None),
                ("mean",                  np.mean, None),
                ("final (steady state)",  None,    HIGHLIGHT_GREEN),
                ("avg last 1 s",          "last1", HIGHLIGHT_YELLOW),
            ]
            for stat_name, fn, hi_color in rows:
                alt = r % 2 == 1
                c1 = ws.cell(row=r, column=1, value=coeff_name)
                c1.border = BORDER
                c2 = ws.cell(row=r, column=2, value=stat_name)
                c2.border = BORDER
                if hi_color:
                    c1.font = Font(bold=True, name="Arial", size=10, color=hi_color)
                    c2.font = Font(bold=True, name="Arial", size=10, color=hi_color)
                else:
                    c1.font = Font(bold=True, name="Arial", size=10)
                    c2.font = CELL_FONT
                col = 3
                for d in ts_data:
                    val = None
                    if d["fc_list"]:
                        fc  = d["fc_list"][0]
                        arr = fc.get(key, np.array([]))
                        if len(arr):
                            if fn is None:
                                val = float(arr[-1])        # final time step
                            elif fn == "last1":
                                val = _last_second_mean(fc, key)
                            else:
                                val = float(fn(arr))
                    cell = _val(ws, r, col, val if val is not None else "N/A", alt)
                    if hi_color and isinstance(val, float):
                        cell.font = Font(bold=True, name="Arial", size=10, color=hi_color)
                    col += 4
                r += 1
            r += 1
        ws.freeze_panes = "C4"

    out = os.path.join(mesh_dir, "stats.xlsx")
    wb.save(out)
    return out


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────

def main():
    root = os.path.abspath(sys.argv[1] if len(sys.argv) > 1 else ".")
    setup_style()

    if not os.path.isdir(root):
        print(f"[ERROR] Not a directory: {root}")
        sys.exit(1)

    if os.path.basename(root).lower() == "postprocessing":
        pp_dirs = [root]
    else:
        pp_dirs = find_postprocessing_dirs(root)

    if not pp_dirs:
        print(f"[ERROR] No 'postProcessing' folders found under: {root}")
        sys.exit(1)

    print(f"[INFO] Root  : {root}")
    print(f"[INFO] Found : {len(pp_dirs)} postProcessing folder(s)\n")

    # Individual plots
    total_fc = total_yp = 0
    for pp_dir in pp_dirs:
        label = relative_label(pp_dir, root)
        print(f"  Individual: {label}")
        fc_s, yp_s = save_individual_plots(pp_dir, root)
        if fc_s:
            print(f"    → forceCoeffs.png  |  forceCoeffs_last1s.png")
        if yp_s:
            print(f"    → yPlus.png  |  yPlus_min/max/mean.png  |  yPlus_*_last1s.png")
        total_fc += fc_s
        total_yp += yp_s

    # Combined mesh-level plots + Excel
    groups = group_by_mesh(pp_dirs)
    print(f"\n[INFO] Found {len(groups)} mesh-level folder(s)\n")

    total_comb = total_xlsx = 0
    for mesh_dir, ts_entries in sorted(groups.items()):
        label = relative_label(mesh_dir, root)
        print(f"  Mesh group : {label}")
        print(f"    Timesteps: {[e[0] for e in ts_entries]}")
        try:
            save_combined_plots(mesh_dir, ts_entries, root)
            print(f"    → forceCoeffs_combined.png  |  forceCoeffs_last1s_combined.png")
            print(f"    → yPlus_combined.png  |  yPlus_*_combined.png  |  yPlus_*_last1s_combined.png")
            total_comb += 1
        except Exception as e:
            print(f"    [WARN] Combined plots failed: {e}")
        try:
            write_excel(mesh_dir, ts_entries, root)
            print(f"    → stats.xlsx")
            total_xlsx += 1
        except Exception as e:
            print(f"    [WARN] Excel export failed: {e}")

    print(f"\n[DONE]")
    print(f"  Individual : {total_fc} forceCoeffs.png  |  {total_yp} yPlus sets")
    print(f"  Combined   : {total_comb} mesh folders  |  {total_xlsx} stats.xlsx files")


if __name__ == "__main__":
    main()
